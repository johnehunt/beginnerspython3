from openpyxl import Workbook


def main():
    print('Starting Write Excel Example with openPyXL')

    workbook = Workbook()
    # Get the current active worksheet
    ws = workbook.active
    ws.title = 'my worksheet'
    ws.sheet_properties.tabColor = '1072BA'

    ws['A1'] = 42
    ws['A2'] = 12
    ws['A3'] = '=SUM(A1, A2)'

    ws2 = workbook.create_sheet(title='my other sheet')
    ws2['A1'] = 3.42
    ws2.append([1, 2, 3])
    ws2.cell(column=2, row=1, value=15)

    workbook.save('sample.xlsx')

    print('Done Write Excel Example')


if __name__ == '__main__':
    main()
