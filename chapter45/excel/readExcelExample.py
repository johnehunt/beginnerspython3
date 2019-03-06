from openpyxl import load_workbook

def main():
    print('Starting reading Excel file using openPyXL')

    wb = load_workbook(filename='sample.xlsx')
    print(wb.active)
    print(wb.sheetnames)
    print(wb.worksheets)

    print('-' * 10)
    ws = wb['my worksheet']
    print(ws['A1'])
    print(ws['A1'].value)
    print(ws['A2'].value)
    print(ws['A3'].value)

    print('-' * 10)
    for sheet in wb:
        print(sheet.title)

    print('-' * 10)
    cell_range = ws['A1':'A3']
    for cell in cell_range:
        print(cell[0].value)
    print('-' * 10)

    print('Finished reading Excel file using openPyXL')

if __name__ == '__main__':
    main()